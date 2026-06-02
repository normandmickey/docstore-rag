import logging
import re
from datetime import timedelta
from io import BytesIO

from django import forms

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import FileResponse, HttpResponse
from django.shortcuts import redirect, render
from django.utils import timezone

from openpyxl import Workbook

from control.views import _dashboard_base, _handle_workspace_actions
from support.models import SupportChannel, SupportConversation, SupportMessage
from integrations.voice.models import VoiceCallRecord
from chatbots.models import ChatbotConversation, ChatbotIntegration
from .models import SpreadsheetTransformJob, SpreadsheetTransformTemplate
from .tasks import build_spreadsheet_transform_export
from .spreadsheet_transform import (
    SpreadsheetTransformError,
    apply_transform_plan,
    build_column_planner,
    build_output_column_planner,
    build_transform_prompt_payload,
    export_transform_csv,
    export_transform_xlsx,
    load_tabular_file,
    plan_transform,
)

logger = logging.getLogger(__name__)


class SpreadsheetTransformForm(forms.Form):
    file = forms.FileField(required=False)
    lookup_file = forms.FileField(required=False, help_text='Optional reference/lookup CSV or XLSX for exact-match lookups.')
    transform_request = forms.CharField(required=False, label='Optional overall notes', widget=forms.Textarea(attrs={'rows': 4}), help_text='Optional extra context for the planner after you define the desired output columns.')
    export_format = forms.ChoiceField(required=False, choices=[('xlsx', 'XLSX'), ('csv', 'CSV')], initial='xlsx')
    strict_sanitization = forms.BooleanField(required=False, initial=False, help_text='Use more aggressive prompt-side masking for sample rows before AI planning.')
    ignore_hidden_rows = forms.BooleanField(required=False, initial=False, help_text='For XLSX files, skip hidden rows during preview and export.')
    ignore_hidden_columns = forms.BooleanField(required=False, initial=False, help_text='For XLSX files, skip hidden columns during preview and export.')
    has_prepared_file = forms.BooleanField(required=False, widget=forms.HiddenInput())


def _infer_support_source(conversation):
    meta = conversation.metadata_json or {}
    source = (meta.get('source') or '').strip().lower()
    if source == 'voice':
        return 'voice'
    if conversation.channel_id:
        channel_type = (conversation.channel.channel_type or '').strip().lower()
        if channel_type == 'voice':
            return 'voice'
        if channel_type == 'sms':
            return 'sms'
        if channel_type == 'both':
            return 'sms/voice'
    return 'support'


def _column_legend(headers: list[str]) -> list[dict]:
    legend = []
    for idx, header in enumerate(headers or []):
        current = idx + 1
        letter = ''
        while current > 0:
            current, remainder = divmod(current - 1, 26)
            letter = chr(65 + remainder) + letter
        legend.append({'letter': letter, 'name': header})
    return legend


def _normalize_header(value: str) -> str:
    return re.sub(r'[^a-z0-9]+', '', (value or '').strip().lower())


def _template_required_headers(template: SpreadsheetTransformTemplate) -> list[str]:
    required = []
    for item in template.output_plan_json or []:
        for key in ['source_a', 'source_b', 'source_hint']:
            value = (item.get(key) or '').strip()
            if value and value not in required:
                required.append(value)
    return required or list(template.source_headers_json or [])


def _template_match_report(expected_headers: list[str], actual_headers: list[str], required_headers: list[str] | None = None) -> dict:
    expected = [header for header in expected_headers or [] if header]
    required_headers = [header for header in (required_headers or []) if header]
    actual = [header for header in actual_headers or [] if header]
    actual_map = {_normalize_header(header): header for header in actual}
    matched = []
    missing = []
    for header in expected:
        normalized = _normalize_header(header)
        if normalized and normalized in actual_map:
            matched.append({'expected': header, 'actual': actual_map[normalized], 'required': header in required_headers})
        else:
            missing.append(header)
    required_missing = [header for header in required_headers if _normalize_header(header) not in actual_map]
    expected_norms = {_normalize_header(header) for header in expected if _normalize_header(header)}
    extra = [header for header in actual if _normalize_header(header) not in expected_norms]
    if expected and not missing and not extra:
        status = 'exact'
    elif not required_missing and matched:
        status = 'usable'
    elif matched:
        status = 'risky'
    else:
        status = 'incompatible'
    return {
        'status': status,
        'matched': matched,
        'missing': missing,
        'required_missing': required_missing,
        'extra': extra,
        'expected_count': len(expected),
        'actual_count': len(actual),
        'can_proceed': status in {'exact', 'usable', 'risky'},
    }


def _read_output_plan_from_post(request) -> list[dict]:
    total = int((request.POST.get('output_plan_total') or '0').strip() or 0)
    items = []
    for idx in range(total):
        items.append({
            'order': int((request.POST.get(f'output_plan_{idx}_order') or str(idx + 1)).strip() or (idx + 1)),
            'name': (request.POST.get(f'output_plan_{idx}_name') or '').strip(),
            'operation': (request.POST.get(f'output_plan_{idx}_operation') or 'keep').strip(),
            'format': (request.POST.get(f'output_plan_{idx}_format') or 'keep_source').strip(),
            'source_a': (request.POST.get(f'output_plan_{idx}_source_a') or '').strip(),
            'source_b': (request.POST.get(f'output_plan_{idx}_source_b') or '').strip(),
            'lookup_key': (request.POST.get(f'output_plan_{idx}_lookup_key') or '').strip(),
            'lookup_return': (request.POST.get(f'output_plan_{idx}_lookup_return') or '').strip(),
            'source_hint': (request.POST.get(f'output_plan_{idx}_source_hint') or '').strip(),
            'instructions': (request.POST.get(f'output_plan_{idx}_instructions') or '').strip(),
        })
    items = [item for item in items if item.get('name') or item.get('source_hint') or item.get('instructions')]
    return sorted(items, key=lambda item: item.get('order') or 0)


def _read_column_plan_from_post(request) -> list[dict]:
    total = int((request.POST.get('column_plan_total') or '0').strip() or 0)
    items = []
    for idx in range(total):
        items.append({
            'source_column': (request.POST.get(f'column_plan_{idx}_source_column') or '').strip(),
            'action': (request.POST.get(f'column_plan_{idx}_action') or 'keep').strip(),
            'target_column': (request.POST.get(f'column_plan_{idx}_target_column') or '').strip(),
            'instructions': (request.POST.get(f'column_plan_{idx}_instructions') or '').strip(),
            'samples': [],
            'detected_sensitive_types': [],
        })
    return [item for item in items if item.get('source_column')]


def spreadsheet_transformer(request):
    base = _dashboard_base(request)
    handled = _handle_workspace_actions(request, base)
    if handled:
        return handled

    base['section'] = 'reports'
    base['report_name'] = 'spreadsheet_transformer'
    base['spreadsheet_transform_form'] = SpreadsheetTransformForm()
    base['spreadsheet_transform_preview_headers'] = []
    base['spreadsheet_transform_preview_rows'] = []
    base['spreadsheet_transform_plan'] = None
    base['spreadsheet_transform_detected_fields'] = {}
    base['spreadsheet_transform_sanitized_samples'] = []
    base['spreadsheet_transform_prompt_preview'] = None
    base['spreadsheet_transform_column_plan'] = []
    base['spreadsheet_transform_source_headers'] = []
    base['spreadsheet_transform_source_rows'] = []
    base['spreadsheet_transform_source_sheet_name'] = ''
    base['spreadsheet_transform_source_row_count'] = 0
    base['spreadsheet_transform_output_plan'] = []
    base['spreadsheet_transform_source_legend'] = []
    base['spreadsheet_transform_lookup_headers'] = []
    base['spreadsheet_transform_lookup_rows'] = []
    base['spreadsheet_transform_lookup_sheet_name'] = ''
    current_tenant = base.get('current_tenant')
    current_workspace = base.get('current_workspace')
    base['spreadsheet_transform_jobs'] = SpreadsheetTransformJob.objects.filter(
        tenant=current_tenant,
        workspace=current_workspace,
    )[:10] if current_tenant and current_workspace else []
    base['spreadsheet_transform_templates'] = SpreadsheetTransformTemplate.objects.filter(
        tenant=current_tenant,
        workspace=current_workspace,
    )[:10] if current_tenant and current_workspace else []
    base['spreadsheet_transform_selected_template_id'] = None
    base['spreadsheet_transform_template_match'] = None

    session_result = request.session.get('spreadsheet_transform_result') or {}
    if session_result:
        base['spreadsheet_transform_form'] = SpreadsheetTransformForm(initial={
            'transform_request': session_result.get('transform_request', ''),
            'export_format': session_result.get('export_format', 'xlsx'),
            'strict_sanitization': session_result.get('strict_sanitization', False),
            'ignore_hidden_rows': session_result.get('ignore_hidden_rows', False),
            'ignore_hidden_columns': session_result.get('ignore_hidden_columns', False),
            'has_prepared_file': True,
        })
        base['spreadsheet_transform_plan'] = session_result.get('plan')
        base['spreadsheet_transform_detected_fields'] = session_result.get('detected_fields', {})
        base['spreadsheet_transform_sanitized_samples'] = session_result.get('sanitized_samples', [])
        base['spreadsheet_transform_prompt_preview'] = session_result.get('prompt_preview')
        base['spreadsheet_transform_column_plan'] = session_result.get('column_plan', [])
        base['spreadsheet_transform_preview_headers'] = session_result.get('headers', [])
        base['spreadsheet_transform_preview_rows'] = (session_result.get('rows') or [])[:20]
        base['spreadsheet_transform_source_headers'] = session_result.get('source_headers', [])
        base['spreadsheet_transform_source_rows'] = session_result.get('source_rows', [])
        base['spreadsheet_transform_source_sheet_name'] = session_result.get('source_sheet_name', '')
        base['spreadsheet_transform_source_row_count'] = session_result.get('source_row_count', 0)
        base['spreadsheet_transform_output_plan'] = session_result.get('output_plan', [])
        base['spreadsheet_transform_source_legend'] = _column_legend(session_result.get('source_headers', []))
        base['spreadsheet_transform_selected_template_id'] = session_result.get('selected_template_id')
        base['spreadsheet_transform_template_match'] = session_result.get('template_match')
        base['spreadsheet_transform_lookup_headers'] = session_result.get('lookup_headers', [])
        base['spreadsheet_transform_lookup_rows'] = session_result.get('lookup_rows', [])
        base['spreadsheet_transform_lookup_sheet_name'] = session_result.get('lookup_sheet_name', '')

    if request.method == 'POST':
        action = (request.POST.get('action') or 'inspect').strip()
        form = SpreadsheetTransformForm(request.POST, request.FILES)
        logger.info('spreadsheet_transformer POST action=%s file_keys=%s post_keys=%s', action, list(request.FILES.keys()), list(request.POST.keys()))
        base['spreadsheet_transform_form'] = form
        if form.is_valid():
            try:
                session_table = request.session.get('spreadsheet_transform_table') or {}
                session_result = request.session.get('spreadsheet_transform_result') or {}

                if action == 'load_template':
                    template_id = (request.POST.get('template_id') or '').strip()
                    current_tenant = base.get('current_tenant')
                    current_workspace = base.get('current_workspace')
                    template = SpreadsheetTransformTemplate.objects.filter(
                        id=template_id,
                        tenant=current_tenant,
                        workspace=current_workspace,
                    ).first()
                    if not template:
                        raise SpreadsheetTransformError('Template not found in this workspace.')
                    existing_result = request.session.get('spreadsheet_transform_result') or {}
                    request.session['spreadsheet_transform_result'] = {
                        'plan': existing_result.get('plan'),
                        'detected_fields': existing_result.get('detected_fields', {}),
                        'sanitized_samples': existing_result.get('sanitized_samples', []),
                        'prompt_preview': existing_result.get('prompt_preview'),
                        'headers': existing_result.get('headers', []),
                        'rows': existing_result.get('rows', []),
                        'export_format': template.export_format or 'xlsx',
                        'strict_sanitization': bool(template.strict_sanitization),
                        'ignore_hidden_rows': bool(template.ignore_hidden_rows),
                        'ignore_hidden_columns': bool(template.ignore_hidden_columns),
                        'transform_request': template.transform_request or '',
                        'column_plan': template.column_plan_json or [],
                        'output_plan': template.output_plan_json or [],
                        'source_headers': existing_result.get('source_headers', template.source_headers_json or []),
                        'source_rows': existing_result.get('source_rows', []),
                        'source_sheet_name': existing_result.get('source_sheet_name', ''),
                        'source_row_count': existing_result.get('source_row_count', 0),
                        'selected_template_id': template.id,
                        'template_match': existing_result.get('template_match'),
                    }
                    request.session.modified = True
                    messages.success(request, f'Loaded template "{template.name}". Now upload a file and inspect it to confirm the input matches.')
                    return redirect('spreadsheet_transformer')

                if action == 'inspect':
                    logger.info('spreadsheet_transformer inspect valid cleaned_has_file=%s', bool(form.cleaned_data.get('file')))
                    if not form.cleaned_data.get('file'):
                        raise SpreadsheetTransformError('Please upload a CSV or XLSX file to inspect.')
                    table = load_tabular_file(form.cleaned_data['file'])
                    logger.info('spreadsheet_transformer inspect parsed sheet=%s rows=%s headers=%s', table.get('sheet_name'), table.get('row_count'), len(table.get('headers') or []))
                    lookup_table = None
                    if form.cleaned_data.get('lookup_file'):
                        lookup_table = load_tabular_file(form.cleaned_data['lookup_file'])
                    request.session['spreadsheet_transform_table'] = table
                    if lookup_table is not None:
                        request.session['spreadsheet_transform_lookup_table'] = lookup_table
                    existing_result = request.session.get('spreadsheet_transform_result') or {}
                    selected_template_id = existing_result.get('selected_template_id')
                    template_match = None
                    output_plan = existing_result.get('output_plan') or build_output_column_planner(table.get('headers') or [])
                    column_plan = existing_result.get('column_plan') or []
                    transform_request = existing_result.get('transform_request', form.cleaned_data['transform_request'])
                    if selected_template_id and current_tenant and current_workspace:
                        template = SpreadsheetTransformTemplate.objects.filter(
                            id=selected_template_id,
                            tenant=current_tenant,
                            workspace=current_workspace,
                        ).first()
                        if template:
                            template_match = _template_match_report(
                                template.source_headers_json or [],
                                table.get('headers') or [],
                                required_headers=_template_required_headers(template),
                            )
                            output_plan = template.output_plan_json or output_plan
                            column_plan = template.column_plan_json or column_plan
                            transform_request = template.transform_request or transform_request
                    request.session['spreadsheet_transform_result'] = {
                        'plan': None,
                        'detected_fields': {},
                        'sanitized_samples': [],
                        'prompt_preview': None,
                        'headers': [],
                        'rows': [],
                        'export_format': form.cleaned_data.get('export_format') or existing_result.get('export_format') or 'xlsx',
                        'strict_sanitization': form.cleaned_data.get('strict_sanitization') or existing_result.get('strict_sanitization') or False,
                        'ignore_hidden_rows': bool(form.cleaned_data.get('ignore_hidden_rows') or existing_result.get('ignore_hidden_rows')),
                        'ignore_hidden_columns': bool(form.cleaned_data.get('ignore_hidden_columns') or existing_result.get('ignore_hidden_columns')),
                        'transform_request': transform_request,
                        'column_plan': column_plan,
                        'output_plan': output_plan,
                        'source_headers': table.get('headers') or [],
                        'source_rows': (table.get('rows') or [])[:20],
                        'source_sheet_name': table.get('sheet_name') or '',
                        'source_row_count': table.get('row_count') or 0,
                        'lookup_headers': (lookup_table or existing_result.get('lookup_headers') and {'headers': existing_result.get('lookup_headers'), 'rows': existing_result.get('lookup_rows', []), 'sheet_name': existing_result.get('lookup_sheet_name', '')} or {}).get('headers', []) if lookup_table is not None else existing_result.get('lookup_headers', []),
                        'lookup_rows': (lookup_table or {}).get('rows', [])[:20] if lookup_table is not None else existing_result.get('lookup_rows', []),
                        'lookup_sheet_name': (lookup_table or {}).get('sheet_name', '') if lookup_table is not None else existing_result.get('lookup_sheet_name', ''),
                        'selected_template_id': selected_template_id,
                        'template_match': template_match,
                    }
                    request.session.modified = True
                    messages.success(request, f"Inspected spreadsheet: {table.get('sheet_name') or 'Sheet1'} with {table.get('row_count') or 0} row(s). Step 2 is ready below.")
                    return redirect('spreadsheet_transformer')

                elif action == 'add_output_column':
                    current = _read_output_plan_from_post(request) or session_result.get('output_plan') or []
                    next_order = (max((item.get('order') or 0) for item in current) + 1) if current else 1
                    current.append({
                        'letter': '',
                        'order': next_order,
                        'name': f'New Column {next_order}',
                        'operation': 'derive',
                        'format': 'keep_source',
                        'source_a': '',
                        'source_b': '',
                        'source_hint': '',
                        'instructions': '',
                    })
                    session_result.update({'output_plan': current})
                    request.session['spreadsheet_transform_result'] = session_result
                    request.session.modified = True
                    messages.success(request, 'Added a new output column to the transformation setup.')
                    return redirect('spreadsheet_transformer')

                elif action == 'prepare':
                    if not session_table:
                        raise SpreadsheetTransformError('No inspected file is available yet. Upload and inspect a CSV or XLSX file first.')
                    table = session_table
                    preliminary_payload, detected_fields, sanitized_samples, _ = build_transform_prompt_payload(
                        headers=table['headers'],
                        rows=table['rows'],
                        user_request=form.cleaned_data['transform_request'],
                        strict_sanitization=form.cleaned_data.get('strict_sanitization') or False,
                    )
                    existing_column_plan = session_result.get('column_plan') or []
                    posted_output_plan = _read_output_plan_from_post(request)
                    existing_output_plan = session_result.get('output_plan') or []
                    column_plan = existing_column_plan or build_column_planner(table['headers'], sanitized_samples, detected_fields)
                    output_plan = posted_output_plan or existing_output_plan or build_output_column_planner(table['headers'])
                    user_payload, detected_fields, sanitized_samples, system_prompt = build_transform_prompt_payload(
                        headers=table['headers'],
                        rows=table['rows'],
                        user_request=form.cleaned_data['transform_request'],
                        strict_sanitization=form.cleaned_data.get('strict_sanitization') or False,
                        column_plan=column_plan,
                        output_plan=output_plan,
                    )
                    request.session['spreadsheet_transform_result'] = {
                        'plan': None,
                        'detected_fields': detected_fields,
                        'sanitized_samples': sanitized_samples,
                        'prompt_preview': {'system_prompt': system_prompt, 'user_payload': user_payload},
                        'headers': [],
                        'rows': [],
                        'export_format': form.cleaned_data.get('export_format') or session_result.get('export_format') or 'xlsx',
                        'strict_sanitization': form.cleaned_data.get('strict_sanitization') or False,
                        'ignore_hidden_rows': bool(form.cleaned_data.get('ignore_hidden_rows') or session_result.get('ignore_hidden_rows')),
                        'ignore_hidden_columns': bool(form.cleaned_data.get('ignore_hidden_columns') or session_result.get('ignore_hidden_columns')),
                        'transform_request': form.cleaned_data['transform_request'],
                        'column_plan': column_plan,
                        'output_plan': output_plan,
                        'source_headers': session_result.get('source_headers', table.get('headers') or []),
                        'source_rows': session_result.get('source_rows', (table.get('rows') or [])[:20]),
                        'source_sheet_name': session_result.get('source_sheet_name', table.get('sheet_name') or ''),
                        'source_row_count': session_result.get('source_row_count', table.get('row_count') or 0),
                    }
                    request.session.modified = True
                    messages.success(request, 'Prepared the planner prompt. Review Step 3 below and then generate the transform preview.')
                    return redirect('spreadsheet_transformer')

                elif action == 'preview':
                    if not session_table:
                        raise SpreadsheetTransformError('No prepared prompt is available yet. Upload a file and prepare the prompt first.')
                    column_plan = _read_column_plan_from_post(request) or session_result.get('column_plan') or []
                    output_plan = _read_output_plan_from_post(request) or session_result.get('output_plan') or []
                    plan, detected_fields, sanitized_samples, prompt_preview = plan_transform(
                        headers=session_table.get('headers') or [],
                        rows=session_table.get('rows') or [],
                        user_request=form.cleaned_data['transform_request'],
                        strict_sanitization=form.cleaned_data.get('strict_sanitization') or False,
                        column_plan=column_plan,
                        output_plan=output_plan,
                    )
                    lookup_table = request.session.get('spreadsheet_transform_lookup_table') or {}
                    headers, transformed_rows = apply_transform_plan(
                        rows=session_table.get('rows') or [],
                        plan=plan,
                        output_plan=output_plan,
                        source_headers=session_table.get('headers') or [],
                        hidden_row_indexes=session_table.get('hidden_row_indexes') or [],
                        hidden_columns=session_table.get('hidden_columns') or [],
                        ignore_hidden_rows=bool(form.cleaned_data.get('ignore_hidden_rows') or session_result.get('ignore_hidden_rows')),
                        ignore_hidden_columns=bool(form.cleaned_data.get('ignore_hidden_columns') or session_result.get('ignore_hidden_columns')),
                        lookup_rows=lookup_table.get('rows') or [],
                    )
                    request.session['spreadsheet_transform_result'] = {
                        'plan': plan,
                        'detected_fields': detected_fields,
                        'sanitized_samples': sanitized_samples,
                        'prompt_preview': prompt_preview,
                        'headers': headers,
                        'rows': transformed_rows,
                        'export_format': form.cleaned_data.get('export_format') or session_result.get('export_format') or 'xlsx',
                        'strict_sanitization': form.cleaned_data.get('strict_sanitization') or False,
                        'ignore_hidden_rows': bool(form.cleaned_data.get('ignore_hidden_rows') or session_result.get('ignore_hidden_rows')),
                        'ignore_hidden_columns': bool(form.cleaned_data.get('ignore_hidden_columns') or session_result.get('ignore_hidden_columns')),
                        'transform_request': form.cleaned_data['transform_request'],
                        'column_plan': column_plan,
                        'output_plan': output_plan,
                        'source_headers': session_result.get('source_headers', session_table.get('headers') or []),
                        'source_rows': session_result.get('source_rows', (session_table.get('rows') or [])[:20]),
                        'source_sheet_name': session_result.get('source_sheet_name', session_table.get('sheet_name') or ''),
                        'source_row_count': session_result.get('source_row_count', session_table.get('row_count') or 0),
                    }
                    request.session.modified = True
                    messages.success(request, 'Previewed the transformed file. Review the preview below or download the transformed file.')
                    return redirect('spreadsheet_transformer')

                elif action == 'clear':
                    request.session.pop('spreadsheet_transform_table', None)
                    request.session.pop('spreadsheet_transform_result', None)
                    request.session.modified = True
                    messages.success(request, 'Cleared the spreadsheet transformer state. You can upload a new file and start over.')
                    return redirect('spreadsheet_transformer')

                elif action == 'save_template':
                    if not session_result:
                        raise SpreadsheetTransformError('No transform setup is available yet. Build a spreadsheet transformation first.')
                    current_tenant = base.get('current_tenant')
                    current_workspace = base.get('current_workspace')
                    if not current_tenant or not current_workspace:
                        raise SpreadsheetTransformError('A tenant and workspace are required to save a template.')
                    template_name = (request.POST.get('template_name') or '').strip() or 'Untitled spreadsheet template'
                    template_description = (request.POST.get('template_description') or '').strip()
                    template_visibility = (request.POST.get('template_visibility') or SpreadsheetTransformTemplate.VISIBILITY_PRIVATE).strip()
                    if template_visibility not in {choice[0] for choice in SpreadsheetTransformTemplate.VISIBILITY_CHOICES}:
                        template_visibility = SpreadsheetTransformTemplate.VISIBILITY_PRIVATE
                    SpreadsheetTransformTemplate.objects.create(
                        tenant=current_tenant,
                        workspace=current_workspace,
                        created_by=request.user,
                        name=template_name,
                        description=template_description,
                        visibility=template_visibility,
                        source_headers_json=session_result.get('source_headers') or [],
                        output_plan_json=session_result.get('output_plan') or [],
                        column_plan_json=session_result.get('column_plan') or [],
                        transform_request=session_result.get('transform_request', ''),
                        export_format=session_result.get('export_format') or 'xlsx',
                        strict_sanitization=bool(session_result.get('strict_sanitization')),
                        ignore_hidden_rows=bool(session_result.get('ignore_hidden_rows')),
                        ignore_hidden_columns=bool(session_result.get('ignore_hidden_columns')),
                    )
                    messages.success(request, f'Saved template "{template_name}".')
                    return redirect('spreadsheet_transformer')

                elif action == 'download':
                    if not session_result:
                        raise SpreadsheetTransformError('No preview is available yet. Run a preview first.')
                    current_tenant = base.get('current_tenant')
                    current_workspace = base.get('current_workspace')
                    if not current_tenant or not current_workspace:
                        raise SpreadsheetTransformError('A tenant and workspace are required to queue the export.')
                    export_format = form.cleaned_data['export_format'] or session_result.get('export_format') or 'xlsx'
                    job = SpreadsheetTransformJob.objects.create(
                        tenant=current_tenant,
                        workspace=current_workspace,
                        created_by=request.user,
                        status=SpreadsheetTransformJob.STATUS_QUEUED,
                        export_format=export_format,
                        source_name=(session_table.get('sheet_name') or session_table.get('source_type') or 'spreadsheet'),
                        transform_request=session_result.get('transform_request', ''),
                        strict_sanitization=bool(session_result.get('strict_sanitization')),
                        plan_json={**(session_result.get('plan') or {}), 'output_plan': session_result.get('output_plan') or []},
                        headers_json=session_result.get('headers') or [],
                        rows_json=session_result.get('rows') or [],
                        ignore_hidden_rows=bool(session_result.get('ignore_hidden_rows')),
                        ignore_hidden_columns=bool(session_result.get('ignore_hidden_columns')),
                    )
                    build_spreadsheet_transform_export.delay(job.id)
                    messages.success(request, f'Queued export job #{job.id}. Refresh to download when ready.')
                    return redirect('spreadsheet_transformer')

                if session_result:
                    base['spreadsheet_transform_plan'] = session_result.get('plan')
                    base['spreadsheet_transform_detected_fields'] = session_result.get('detected_fields', {})
                    base['spreadsheet_transform_sanitized_samples'] = session_result.get('sanitized_samples', [])
                    base['spreadsheet_transform_prompt_preview'] = session_result.get('prompt_preview')
                    base['spreadsheet_transform_column_plan'] = session_result.get('column_plan', [])
                    base['spreadsheet_transform_preview_headers'] = session_result.get('headers', [])
                    base['spreadsheet_transform_preview_rows'] = (session_result.get('rows') or [])[:20]
                    base['spreadsheet_transform_source_headers'] = session_result.get('source_headers', [])
                    base['spreadsheet_transform_source_rows'] = session_result.get('source_rows', [])
                    base['spreadsheet_transform_source_sheet_name'] = session_result.get('source_sheet_name', '')
                    base['spreadsheet_transform_source_row_count'] = session_result.get('source_row_count', 0)
            except SpreadsheetTransformError as exc:
                logger.warning('spreadsheet_transformer handled error action=%s error=%s', action, exc)
                messages.error(request, str(exc))
            except Exception as exc:
                logger.exception('spreadsheet_transformer unexpected failure action=%s', action)
                messages.error(request, f'Spreadsheet transform failed: {exc}')
        else:
            logger.warning('spreadsheet_transformer invalid form action=%s errors=%s', action, form.errors.as_json())

    return render(request, 'dashboard/spreadsheet_transformer.html', base)


@login_required
def spreadsheet_transform_download(request, job_id: int):
    base = _dashboard_base(request)
    handled = _handle_workspace_actions(request, base)
    if handled:
        return handled
    current_tenant = base.get('current_tenant')
    current_workspace = base.get('current_workspace')
    job = SpreadsheetTransformJob.objects.filter(
        id=job_id,
        tenant=current_tenant,
        workspace=current_workspace,
        created_by=request.user,
    ).first()
    if not job or not job.output_file:
        messages.error(request, 'Export file is not ready yet.')
        return redirect('spreadsheet_transformer')
    return FileResponse(job.output_file.open('rb'), as_attachment=True, filename=job.output_file.name.rsplit('/', 1)[-1])


@login_required
def support_activity_report(request):
    base = _dashboard_base(request)
    handled = _handle_workspace_actions(request, base)
    if handled:
        return handled

    tenant = base.get('current_tenant')
    if tenant is None:
        messages.error(request, 'No tenant selected.')
        return redirect('dashboard')

    today = timezone.localdate()
    start_default = today - timedelta(days=29)
    start_raw = (request.GET.get('start') or '').strip()
    end_raw = (request.GET.get('end') or '').strip()

    try:
        start_date = timezone.datetime.fromisoformat(start_raw).date() if start_raw else start_default
    except ValueError:
        start_date = start_default
    try:
        end_date = timezone.datetime.fromisoformat(end_raw).date() if end_raw else today
    except ValueError:
        end_date = today

    if end_date < start_date:
        start_date, end_date = end_date, start_date

    start_dt = timezone.make_aware(timezone.datetime.combine(start_date, timezone.datetime.min.time()))
    end_dt = timezone.make_aware(timezone.datetime.combine(end_date, timezone.datetime.max.time()))

    channel_id_raw = (request.GET.get('channel') or '').strip()
    assigned_user_id_raw = (request.GET.get('assigned_user') or '').strip()
    channel_id = int(channel_id_raw) if channel_id_raw.isdigit() else None
    assigned_user_id = int(assigned_user_id_raw) if assigned_user_id_raw.isdigit() else None

    conversations = SupportConversation.objects.filter(tenant=tenant, created_at__gte=start_dt, created_at__lte=end_dt)
    messages_qs = SupportMessage.objects.filter(conversation__tenant=tenant, created_at__gte=start_dt, created_at__lte=end_dt)
    calls_qs = VoiceCallRecord.objects.filter(tenant=tenant, created_at__gte=start_dt, created_at__lte=end_dt)

    if channel_id:
        conversations = conversations.filter(channel_id=channel_id)
        messages_qs = messages_qs.filter(conversation__channel_id=channel_id)
        calls_qs = calls_qs.filter(support_channel_id=channel_id)
    if assigned_user_id:
        conversations = conversations.filter(assigned_user_id=assigned_user_id)
        messages_qs = messages_qs.filter(conversation__assigned_user_id=assigned_user_id)
        calls_qs = calls_qs.filter(related_conversation__assigned_user_id=assigned_user_id)

    rows = []
    cursor = start_date
    while cursor <= end_date:
        next_day = cursor + timedelta(days=1)
        day_start = timezone.make_aware(timezone.datetime.combine(cursor, timezone.datetime.min.time()))
        day_end = timezone.make_aware(timezone.datetime.combine(cursor, timezone.datetime.max.time()))
        day_conversations = conversations.filter(created_at__gte=day_start, created_at__lte=day_end)
        rows.append({
            'date': cursor.isoformat(),
            'conversations_created': day_conversations.count(),
            'open_count': day_conversations.filter(status=SupportConversation.STATUS_OPEN).count(),
            'pending_count': day_conversations.filter(status=SupportConversation.STATUS_PENDING).count(),
            'closed_count': day_conversations.filter(status=SupportConversation.STATUS_CLOSED).count(),
            'messages_inbound': messages_qs.filter(created_at__gte=day_start, created_at__lte=day_end, direction=SupportMessage.DIR_INBOUND).count(),
            'messages_outbound': messages_qs.filter(created_at__gte=day_start, created_at__lte=day_end, direction=SupportMessage.DIR_OUTBOUND).count(),
            'voice_calls': calls_qs.filter(created_at__gte=day_start, created_at__lte=day_end).count(),
        })
        cursor = next_day

    summary = {
        'conversations_created': conversations.count(),
        'messages_inbound': messages_qs.filter(direction=SupportMessage.DIR_INBOUND).count(),
        'messages_outbound': messages_qs.filter(direction=SupportMessage.DIR_OUTBOUND).count(),
        'voice_calls': calls_qs.count(),
        'currently_open': SupportConversation.objects.filter(tenant=tenant, status=SupportConversation.STATUS_OPEN).count(),
        'currently_pending': SupportConversation.objects.filter(tenant=tenant, status=SupportConversation.STATUS_PENDING).count(),
    }

    detail_rows = []
    conversation_list = list(conversations.select_related('channel', 'contact', 'assigned_user').order_by('-created_at')[:250])
    for conversation in conversation_list:
        contact = conversation.contact
        contact_meta = contact.metadata_json or {}
        detail_rows.append({
            'created_at': timezone.localtime(conversation.created_at).strftime('%Y-%m-%d %H:%M'),
            'source': _infer_support_source(conversation),
            'contact_name': contact.name or '',
            'contact_phone': contact.phone_number or '',
            'contact_email': (contact_meta.get('email') or '').strip(),
            'channel_name': conversation.channel.name if conversation.channel_id else '',
            'assigned_user': conversation.assigned_user.username if conversation.assigned_user_id else '',
            'subject': conversation.subject or '',
            'status': conversation.status,
            'last_message_at': timezone.localtime(conversation.last_message_at).strftime('%Y-%m-%d %H:%M') if conversation.last_message_at else '',
            'conversation_id': conversation.id,
        })

    if (request.GET.get('format') or '').strip().lower() == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Support Activity'
        ws.append([
            'Date',
            'Conversations Created',
            'Open',
            'Pending',
            'Closed',
            'Inbound Messages',
            'Outbound Messages',
            'Voice Calls',
        ])
        for row in rows:
            ws.append([
                row['date'],
                row['conversations_created'],
                row['open_count'],
                row['pending_count'],
                row['closed_count'],
                row['messages_inbound'],
                row['messages_outbound'],
                row['voice_calls'],
            ])
        ws.append([])
        ws.append(['Summary'])
        for key, value in summary.items():
            ws.append([key, value])

        details = wb.create_sheet(title='Conversation Details')
        details.append([
            'Created At',
            'Source',
            'Contact Name',
            'Contact Phone',
            'Contact Email',
            'Channel',
            'Assigned User',
            'Subject',
            'Status',
            'Last Message At',
            'Conversation ID',
        ])
        for row in detail_rows:
            details.append([
                row['created_at'],
                row['source'],
                row['contact_name'],
                row['contact_phone'],
                row['contact_email'],
                row['channel_name'],
                row['assigned_user'],
                row['subject'],
                row['status'],
                row['last_message_at'],
                row['conversation_id'],
            ])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="support-activity-{start_date.isoformat()}-to-{end_date.isoformat()}.xlsx"'
        return response

    base.update({
        'section': 'reports',
        'report_name': 'support_activity',
        'report_start': start_date.isoformat(),
        'report_end': end_date.isoformat(),
        'report_channel_id': channel_id_raw,
        'report_assigned_user_id': assigned_user_id_raw,
        'report_channels': SupportChannel.objects.filter(tenant=tenant).order_by('name'),
        'report_assigned_users': tenant.memberships.select_related('user').order_by('user__username'),
        'report_rows': rows,
        'report_summary': summary,
        'report_detail_rows': detail_rows,
    })
    return render(request, 'dashboard/report_support_activity.html', base)


@login_required
def interactions_report(request):
    base = _dashboard_base(request)
    handled = _handle_workspace_actions(request, base)
    if handled:
        return handled

    tenant = base.get('current_tenant')
    if tenant is None:
        messages.error(request, 'No tenant selected.')
        return redirect('dashboard')

    today = timezone.localdate()
    start_default = today - timedelta(days=29)
    start_raw = (request.GET.get('start') or '').strip()
    end_raw = (request.GET.get('end') or '').strip()
    source_filter = (request.GET.get('source') or '').strip().lower()
    workspace_id_raw = (request.GET.get('workspace') or '').strip()
    workspace_id = int(workspace_id_raw) if workspace_id_raw.isdigit() else None

    try:
        start_date = timezone.datetime.fromisoformat(start_raw).date() if start_raw else start_default
    except ValueError:
        start_date = start_default
    try:
        end_date = timezone.datetime.fromisoformat(end_raw).date() if end_raw else today
    except ValueError:
        end_date = today

    if end_date < start_date:
        start_date, end_date = end_date, start_date

    start_dt = timezone.make_aware(timezone.datetime.combine(start_date, timezone.datetime.min.time()))
    end_dt = timezone.make_aware(timezone.datetime.combine(end_date, timezone.datetime.max.time()))

    rows = []

    support_qs = SupportConversation.objects.select_related('channel', 'contact', 'assigned_user', 'workspace_context').filter(
        tenant=tenant,
        created_at__gte=start_dt,
        created_at__lte=end_dt,
    )
    if workspace_id:
        support_qs = support_qs.filter(workspace_context_id=workspace_id)
    for conversation in support_qs[:250]:
        source = _infer_support_source(conversation)
        if source_filter and source_filter != source:
            continue
        contact_meta = conversation.contact.metadata_json or {}
        rows.append({
            'timestamp': conversation.created_at,
            'source': source,
            'end_user': conversation.contact.name or conversation.contact.phone_number or '',
            'contact_detail': (contact_meta.get('email') or '').strip() or conversation.contact.phone_number or '',
            'channel': conversation.channel.name if conversation.channel_id else '',
            'workspace': conversation.workspace_context.name if conversation.workspace_context_id else '',
            'assigned_user': conversation.assigned_user.username if conversation.assigned_user_id else '',
            'status': conversation.status,
            'preview': conversation.subject or '',
            'reference': f'support-{conversation.id}',
        })

    voice_qs = VoiceCallRecord.objects.select_related('workspace', 'support_channel').filter(
        tenant=tenant,
        created_at__gte=start_dt,
        created_at__lte=end_dt,
    )
    if workspace_id:
        voice_qs = voice_qs.filter(workspace_id=workspace_id)
    for call in voice_qs[:250]:
        source = 'voice'
        if source_filter and source_filter != source:
            continue
        rows.append({
            'timestamp': call.created_at,
            'source': source,
            'end_user': call.from_number or '',
            'contact_detail': call.to_number or '',
            'channel': call.support_channel.name if call.support_channel_id else '',
            'workspace': call.workspace.name if call.workspace_id else '',
            'assigned_user': '',
            'status': call.close_reason or '',
            'preview': f'Call {call.call_sid}',
            'reference': f'voice-{call.id}',
        })

    chatbot_qs = ChatbotConversation.objects.select_related('integration', 'endpoint', 'workspace').filter(
        tenant=tenant,
        created_at__gte=start_dt,
        created_at__lte=end_dt,
    )
    if workspace_id:
        chatbot_qs = chatbot_qs.filter(workspace_id=workspace_id)
    for convo in chatbot_qs[:250]:
        source = f'chatbot_{convo.platform}'
        if source_filter and source_filter != source:
            continue
        meta = convo.metadata_json or {}
        external_user = meta.get('external_user_name') or meta.get('external_user_id') or convo.external_conversation_id or ''
        channel_label = ''
        if convo.integration_id:
            channel_label = f'{convo.integration.name} ({convo.platform})'
        elif convo.platform:
            channel_label = convo.platform
        rows.append({
            'timestamp': convo.created_at,
            'source': source,
            'end_user': external_user,
            'contact_detail': convo.external_thread_id or convo.external_conversation_id or '',
            'channel': channel_label,
            'workspace': convo.workspace.name if convo.workspace_id else '',
            'assigned_user': '',
            'status': convo.status,
            'preview': convo.title or '',
            'reference': f'chatbot-{convo.id}',
        })

    rows.sort(key=lambda item: item['timestamp'], reverse=True)
    rows = rows[:500]

    summary = {
        'total_rows': len(rows),
        'support_rows': sum(1 for row in rows if row['source'] in {'support', 'sms', 'sms/voice', 'voice'} and row['reference'].startswith('support-')),
        'voice_rows': sum(1 for row in rows if row['reference'].startswith('voice-')),
        'chatbot_rows': sum(1 for row in rows if row['reference'].startswith('chatbot-')),
    }

    if (request.GET.get('format') or '').strip().lower() == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Interactions'
        ws.append(['Timestamp', 'Source', 'End User', 'Contact Detail', 'Channel', 'Workspace', 'Assigned User', 'Status', 'Preview', 'Reference'])
        for row in rows:
            ws.append([
                timezone.localtime(row['timestamp']).strftime('%Y-%m-%d %H:%M') if row['timestamp'] else '',
                row['source'],
                row['end_user'],
                row['contact_detail'],
                row['channel'],
                row['workspace'],
                row['assigned_user'],
                row['status'],
                row['preview'],
                row['reference'],
            ])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="interactions-{start_date.isoformat()}-to-{end_date.isoformat()}.xlsx"'
        return response

    base.update({
        'section': 'reports',
        'report_name': 'interactions',
        'report_start': start_date.isoformat(),
        'report_end': end_date.isoformat(),
        'report_source': source_filter,
        'report_workspace_id': workspace_id_raw,
        'report_source_choices': [
            ('', 'All sources'),
            ('support', 'Support'),
            ('sms', 'SMS'),
            ('sms/voice', 'SMS/Voice'),
            ('voice', 'Voice'),
            (f'chatbot_{ChatbotIntegration.PLATFORM_TELEGRAM}', 'Chatbot · Telegram'),
            (f'chatbot_{ChatbotIntegration.PLATFORM_DISCORD}', 'Chatbot · Discord'),
        ],
        'report_workspaces': tenant.workspaces.order_by('name'),
        'report_rows': rows,
        'report_summary': summary,
    })
    return render(request, 'dashboard/report_interactions.html', base)
