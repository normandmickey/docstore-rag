from __future__ import annotations

import csv
import json
import re
from io import BytesIO, StringIO
from typing import Any

from django.conf import settings
from faker import Faker
from openai import OpenAI
from openpyxl import Workbook, load_workbook


class SpreadsheetTransformError(Exception):
    pass


def _extract_json_object(text: str) -> str:
    value = (text or '').strip()
    if not value:
        raise SpreadsheetTransformError('No transform plan was returned.')
    fenced = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', value, re.DOTALL | re.IGNORECASE)
    if fenced:
        return fenced.group(1).strip()
    start = value.find('{')
    end = value.rfind('}')
    if start >= 0 and end > start:
        return value[start:end + 1].strip()
    return value


def load_tabular_file(uploaded_file) -> dict[str, Any]:
    name = (getattr(uploaded_file, 'name', '') or '').lower()
    raw = uploaded_file.read()
    uploaded_file.seek(0)

    if name.endswith('.csv'):
        text = raw.decode('utf-8-sig', errors='ignore')
        reader = csv.DictReader(StringIO(text))
        rows = [dict(row) for row in reader]
        headers = list(reader.fieldnames or [])
        return {
            'sheet_name': 'Sheet1',
            'headers': headers,
            'rows': rows,
            'row_count': len(rows),
            'source_type': 'csv',
        }

    if name.endswith('.xlsx'):
        wb = load_workbook(BytesIO(raw), data_only=True)
        ws = wb[wb.sheetnames[0]]
        values = list(ws.iter_rows(values_only=True))
        if not values:
            return {
                'sheet_name': ws.title,
                'headers': [],
                'rows': [],
                'row_count': 0,
                'source_type': 'xlsx',
            }
        raw_headers = [str(value).strip() if value is not None else '' for value in values[0]]
        headers = []
        seen = {}
        for idx, header in enumerate(raw_headers):
            key = header or f'column_{idx + 1}'
            if key in seen:
                seen[key] += 1
                key = f'{key}_{seen[key]}'
            else:
                seen[key] = 1
            headers.append(key)
        rows = []
        for value_row in values[1:]:
            row = {}
            for idx, header in enumerate(headers):
                key = header
                cell = value_row[idx] if idx < len(value_row) else None
                row[key] = '' if cell is None else str(cell)
            if any(str(v).strip() for v in row.values()):
                rows.append(row)
        hidden_row_indexes = [idx - 2 for idx in range(2, ws.max_row + 1) if ws.row_dimensions[idx].hidden]
        hidden_columns = []
        for idx, header in enumerate(headers, start=1):
            column_letter = ws.cell(row=1, column=idx).column_letter
            if ws.column_dimensions[column_letter].hidden:
                hidden_columns.append(header)
        return {
            'sheet_name': ws.title,
            'headers': headers,
            'rows': rows,
            'row_count': len(rows),
            'source_type': 'xlsx',
            'hidden_row_indexes': hidden_row_indexes,
            'hidden_columns': hidden_columns,
        }

    raise SpreadsheetTransformError('Unsupported file type. Please upload a CSV or XLSX file.')


def _sample_rows(rows: list[dict[str, Any]], limit: int = 8) -> list[dict[str, Any]]:
    return rows[:limit]


def _looks_like_email(value: str) -> bool:
    return '@' in value and '.' in value.split('@')[-1]


def _looks_like_phone(value: str) -> bool:
    digits = ''.join(ch for ch in value if ch.isdigit())
    return len(digits) >= 10 and len(digits) <= 15


def _looks_like_person_name(header: str, value: str) -> bool:
    header_lower = (header or '').lower()
    if 'name' not in header_lower:
        return False
    parts = [part for part in value.strip().split() if part]
    return 1 <= len(parts) <= 4 and all(part[:1].isalpha() for part in parts)


def _looks_like_ssn(header: str, value: str) -> bool:
    header_lower = (header or '').lower()
    digits = ''.join(ch for ch in value if ch.isdigit())
    if 'ssn' in header_lower or 'social security' in header_lower:
        return len(digits) == 9
    return bool(__import__('re').fullmatch(r'\d{3}-?\d{2}-?\d{4}', value.strip()))


def _looks_like_bank_info(header: str, value: str) -> bool:
    header_lower = (header or '').lower()
    digits = ''.join(ch for ch in value if ch.isdigit())
    bank_markers = ['bank', 'routing', 'account', 'acct', 'iban']
    if any(marker in header_lower for marker in bank_markers):
        return len(digits) >= 6 or len(value.strip()) >= 8
    return False


def _looks_like_dob(header: str, value: str) -> bool:
    header_lower = (header or '').lower()
    if not any(marker in header_lower for marker in ['dob', 'birth', 'date of birth']):
        return False
    value = value.strip()
    return any(ch.isdigit() for ch in value) and any(sep in value for sep in ['/', '-', '.'])


def _looks_like_member_id(header: str, value: str) -> bool:
    header_lower = (header or '').lower()
    id_markers = ['employee id', 'employeeid', 'member id', 'memberid', 'subscriber id', 'subscriberid', 'person id', 'group number', 'group no', 'policy number', 'policy no']
    if any(marker in header_lower for marker in id_markers):
        compact = ''.join(ch for ch in value if ch.isalnum())
        return len(compact) >= 5
    return False


def _looks_like_tax_id(header: str, value: str) -> bool:
    header_lower = (header or '').lower()
    digits = ''.join(ch for ch in value if ch.isdigit())
    tax_markers = ['ein', 'tax id', 'taxid', 'tin', 'federal id']
    if any(marker in header_lower for marker in tax_markers):
        return len(digits) == 9
    return False


def _looks_like_license_or_passport(header: str, value: str) -> bool:
    header_lower = (header or '').lower()
    markers = ['driver', 'license', 'licence', 'passport']
    if any(marker in header_lower for marker in markers):
        compact = ''.join(ch for ch in value if ch.isalnum())
        return len(compact) >= 6
    return False


def _looks_like_address(header: str, value: str) -> bool:
    header_lower = (header or '').lower()
    return any(marker in header_lower for marker in ['address', 'street', 'city', 'state', 'zip', 'postal'])


def sanitize_sample_rows(headers: list[str], rows: list[dict[str, Any]], strict: bool = False) -> tuple[list[dict[str, Any]], dict[str, list[str]]]:
    fake = Faker()
    fake.seed_instance(42)
    replacements = {}
    detected = {}

    def mark(header: str, kind: str):
        detected.setdefault(header, [])
        if kind not in detected[header]:
            detected[header].append(kind)

    def replace(kind: str, original: str):
        key = (kind, original)
        if key in replacements:
            return replacements[key]
        if kind == 'email':
            value = fake.email()
        elif kind == 'phone':
            value = fake.phone_number()
        elif kind == 'name':
            value = fake.name()
        elif kind == 'address':
            value = fake.street_address()
        elif kind == 'city':
            value = fake.city()
        elif kind == 'state':
            value = fake.state_abbr()
        elif kind == 'zip':
            value = fake.postcode()
        elif kind == 'ssn':
            value = fake.ssn()
        elif kind == 'routing':
            value = ''.join(str(fake.random_digit()) for _ in range(9))
        elif kind == 'account':
            value = ''.join(str(fake.random_digit()) for _ in range(12))
        elif kind == 'dob':
            value = fake.date_of_birth(minimum_age=21, maximum_age=70).strftime('%m/%d/%Y')
        elif kind == 'member_id':
            value = ''.join(str(fake.random_digit()) for _ in range(8))
        elif kind == 'tax_id':
            value = ''.join(str(fake.random_digit()) for _ in range(9))
        elif kind == 'license_or_passport':
            value = ''.join(fake.random_choices(elements='ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789', length=9))
        else:
            value = original
        replacements[key] = value
        return value

    sanitized = []
    for row in rows:
        out = {}
        for header in headers:
            value = str(row.get(header, '') or '')
            header_lower = (header or '').lower()
            if not value:
                out[header] = value
            elif _looks_like_email(value):
                mark(header, 'email')
                out[header] = replace('email', value)
            elif _looks_like_phone(value):
                mark(header, 'phone')
                out[header] = replace('phone', value)
            elif _looks_like_person_name(header, value):
                mark(header, 'name')
                out[header] = replace('name', value)
            elif _looks_like_ssn(header, value):
                mark(header, 'ssn')
                out[header] = replace('ssn', value)
            elif _looks_like_bank_info(header, value):
                if 'routing' in header_lower:
                    mark(header, 'routing')
                    out[header] = replace('routing', value)
                else:
                    mark(header, 'account')
                    out[header] = replace('account', value)
            elif _looks_like_dob(header, value):
                mark(header, 'dob')
                out[header] = replace('dob', value)
            elif _looks_like_member_id(header, value):
                mark(header, 'member_id')
                out[header] = replace('member_id', value)
            elif _looks_like_tax_id(header, value):
                mark(header, 'tax_id')
                out[header] = replace('tax_id', value)
            elif _looks_like_license_or_passport(header, value):
                mark(header, 'license_or_passport')
                out[header] = replace('license_or_passport', value)
            elif _looks_like_address(header, value):
                if 'city' in header_lower:
                    mark(header, 'city')
                    out[header] = replace('city', value)
                elif 'state' in header_lower:
                    mark(header, 'state')
                    out[header] = replace('state', value)
                elif 'zip' in header_lower or 'postal' in header_lower:
                    mark(header, 'zip')
                    out[header] = replace('zip', value)
                else:
                    mark(header, 'address')
                    out[header] = replace('address', value)
            else:
                if strict and value and len(value.strip()) >= 6 and any(ch.isdigit() for ch in value):
                    mark(header, 'strict_generic_numeric')
                    out[header] = ''.join(str(fake.random_digit()) if ch.isdigit() else ch for ch in value)
                else:
                    out[header] = value
        sanitized.append(out)
    return sanitized, detected


def _column_letter(index: int) -> str:
    result = ''
    current = index + 1
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        result = chr(65 + remainder) + result
    return result


def build_output_column_planner(headers: list[str]) -> list[dict[str, Any]]:
    return [
        {
            'letter': _column_letter(idx),
            'order': idx + 1,
            'name': header,
            'operation': 'keep',
            'format': 'keep_source',
            'source_a': header,
            'source_b': '',
            'lookup_key': '',
            'lookup_return': '',
            'source_hint': header,
            'instructions': '',
        }
        for idx, header in enumerate(headers)
    ]


def build_column_planner(headers: list[str], rows: list[dict[str, Any]], detected: dict[str, list[str]] | None = None) -> list[dict[str, Any]]:
    detected = detected or {}
    sample_rows = _sample_rows(rows, limit=3)
    planner = []
    for header in headers:
        samples = []
        for row in sample_rows:
            value = str((row or {}).get(header, '') or '').strip()
            if value:
                samples.append(value)
        planner.append({
            'letter': _column_letter(len(planner)),
            'source_column': header,
            'samples': samples[:3],
            'detected_sensitive_types': detected.get(header, []),
            'action': 'keep',
            'target_column': header,
            'instructions': '',
        })
    return planner


def build_transform_prompt_payload(*, headers: list[str], rows: list[dict[str, Any]], user_request: str, strict_sanitization: bool = False, column_plan: list[dict[str, Any]] | None = None, output_plan: list[dict[str, Any]] | None = None) -> tuple[dict[str, Any], dict[str, list[str]], list[dict[str, Any]], str]:
    prompt = (
        'You are planning a spreadsheet transformation. '
        'Return valid JSON only. '
        'Given source headers, sample rows, and a user request, produce a transform plan with this shape: '
        '{"output_columns": [{"name": string, "source": string|null, "instruction": string}], '
        '"filters": [{"column": string, "operator": string, "value": string}], '
        '"notes": [string]}. '
        'Use source=null only when the column must be derived from multiple fields or a transformation instruction. '
        'Be conservative. Do not invent source columns that are not present. '
        'Supported filter operators: equals, contains, not_equals. '
        'If the request is broad, still propose the most likely output columns. '
    )
    sanitized_rows, detected = sanitize_sample_rows(headers, _sample_rows(rows), strict=strict_sanitization)
    planner = column_plan or build_column_planner(headers, sanitized_rows, detected)
    desired_output_plan = output_plan or build_output_column_planner(headers)
    user_input = {
        'source_headers': headers,
        'sample_rows': sanitized_rows,
        'user_request': user_request,
        'column_plan': planner,
        'desired_output_plan': desired_output_plan,
        'sample_rows_are_sanitized': True,
        'strict_sanitization': strict_sanitization,
        'instruction': 'Sample row values may be privacy-sanitized, but headers and structure reflect the real file. Build the transform plan against the real source headers. Use the column_plan as the preferred field-by-field instruction layer when present.',
    }
    return user_input, detected, sanitized_rows, prompt


def plan_transform(*, headers: list[str], rows: list[dict[str, Any]], user_request: str, strict_sanitization: bool = False, column_plan: list[dict[str, Any]] | None = None, output_plan: list[dict[str, Any]] | None = None) -> tuple[dict[str, Any], dict[str, list[str]], list[dict[str, Any]], dict[str, Any]]:
    if not settings.OPENAI_API_KEY:
        raise SpreadsheetTransformError('OPENAI_API_KEY is not configured.')

    client = OpenAI(api_key=settings.OPENAI_API_KEY)
    user_input, detected, sanitized_rows, prompt = build_transform_prompt_payload(
        headers=headers,
        rows=rows,
        user_request=user_request,
        strict_sanitization=strict_sanitization,
        column_plan=column_plan,
        output_plan=output_plan,
    )
    response = client.responses.create(
        model=getattr(settings, 'SPREADSHEET_TRANSFORM_MODEL', 'gpt-4.1-mini'),
        input=[
            {'role': 'system', 'content': [{'type': 'input_text', 'text': prompt}]},
            {'role': 'user', 'content': [{'type': 'input_text', 'text': json.dumps(user_input)}]},
        ],
    )
    text = (getattr(response, 'output_text', '') or '').strip()
    if not text:
        raise SpreadsheetTransformError('No transform plan was returned.')
    try:
        plan = json.loads(_extract_json_object(text))
        return plan, detected, sanitized_rows, {'system_prompt': prompt, 'user_payload': user_input}
    except json.JSONDecodeError as exc:
        raise SpreadsheetTransformError(f'Could not parse transform plan JSON: {exc}') from exc


def _coerce_numeric(value: Any) -> float | None:
    text = str(value or '').strip()
    if not text:
        return None
    negative = False
    if text.startswith('(') and text.endswith(')'):
        negative = True
        text = text[1:-1]
    text = text.replace('$', '').replace(',', '').replace('%', '').strip()
    try:
        number = float(text)
    except ValueError:
        return None
    return -number if negative else number


def _format_numeric_like_source(value: float, source_value: Any = '') -> str:
    source_text = str(source_value or '').strip()
    if '%' in source_text:
        return f'{value:.2f}%'
    if source_text.startswith('$'):
        return f'${value:.2f}'
    if float(value).is_integer():
        return str(int(value))
    return f'{value:.2f}'


def _resolve_source_name(source: str, row: dict[str, Any]) -> str:
    source = (source or '').strip()
    if not source:
        return ''
    if source in row:
        return source
    match = re.fullmatch(r'([A-Z]+)', source.upper())
    if match:
        letters = match.group(1)
        index = 0
        for ch in letters:
            index = index * 26 + (ord(ch) - 64)
        index -= 1
        keys = list(row.keys())
        if 0 <= index < len(keys):
            return keys[index]
    return source


def _parse_hours_minutes(value: Any) -> float | None:
    text = str(value or '').strip()
    if not text:
        return None
    match = re.fullmatch(r'(\d+):(\d{1,2})', text)
    if not match:
        return None
    hours = int(match.group(1))
    minutes = int(match.group(2))
    return hours + (minutes / 60.0)


def _filter_visible_rows(rows: list[dict[str, Any]], output_plan: list[dict[str, Any]] | None) -> list[dict[str, Any]]:
    if not rows or not output_plan:
        return rows
    wants_visible_only = any(
        'visible' in (item.get('instructions') or '').lower() or (item.get('operation') or '').lower() == 'visible_only'
        for item in output_plan
    )
    if not wants_visible_only:
        return rows
    visibility_keys = [key for key in rows[0].keys() if 'visible' in key.lower() or 'hidden' in key.lower()]
    if not visibility_keys:
        return rows
    filtered = []
    for row in rows:
        keep = True
        for key in visibility_keys:
            value = str(row.get(key, '') or '').strip().lower()
            if 'hidden' in key.lower() and value in {'true', 'yes', '1', 'hidden'}:
                keep = False
            elif 'visible' in key.lower() and value not in {'true', 'yes', '1', 'visible', 'shown'}:
                keep = False
        if keep:
            filtered.append(row)
    return filtered


def _apply_output_plan_lookup(rows: list[dict[str, Any]], output_plan: list[dict[str, Any]] | None, lookup_rows: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    if not output_plan or not lookup_rows:
        return rows
    transformed_rows = [dict(row) for row in rows]
    for item in output_plan:
        target = (item.get('name') or '').strip()
        operation = (item.get('operation') or '').strip().lower()
        if operation != 'lookup' or not target:
            continue
        primary_key = (item.get('source_a') or '').strip()
        lookup_key = (item.get('lookup_key') or '').strip()
        lookup_return = (item.get('lookup_return') or '').strip()
        if not primary_key or not lookup_key or not lookup_return:
            continue
        lookup_index = {}
        for lookup_row in lookup_rows:
            lookup_value = str(lookup_row.get(lookup_key, '') or '').strip()
            if lookup_value:
                lookup_index[lookup_value] = lookup_row.get(lookup_return, '')
        for row in transformed_rows:
            match_value = str(row.get(primary_key, '') or '').strip()
            if match_value in lookup_index:
                row[target] = lookup_index[match_value]
    return transformed_rows


def _apply_output_plan_arithmetic(rows: list[dict[str, Any]], output_plan: list[dict[str, Any]] | None) -> list[dict[str, Any]]:
    if not output_plan:
        return rows
    transformed_rows = [dict(row) for row in rows]
    for item in output_plan:
        target = (item.get('name') or '').strip()
        operation = (item.get('operation') or '').strip().lower()
        source_hint = (item.get('source_hint') or '').strip()
        instruction = (item.get('instructions') or '').strip().lower()
        if not target or target not in (rows[0].keys() if rows else []):
            continue
        if operation == 'time_to_decimal_hours':
            source_name = _resolve_source_name(item.get('source_a') or target, rows[0] if rows else {})
            if rows and source_name not in rows[0]:
                continue
            for row in transformed_rows:
                converted = _parse_hours_minutes(row.get(source_name, ''))
                if converted is None:
                    continue
                row[target] = _format_numeric_like_source(converted, converted)
            continue
        if operation not in {'multiply', 'divide', 'add', 'subtract'} and all(token not in instruction for token in ['multiply', 'times', '*', 'divide', '/', 'add', '+', 'subtract', '-']):
            continue
        left_source = _resolve_source_name(item.get('source_a') or '', rows[0] if rows else {})
        right_source = _resolve_source_name(item.get('source_b') or '', rows[0] if rows else {})
        if not left_source or not right_source:
            source_candidates = [part.strip() for part in re.split(r'[,+]| and ', source_hint) if part.strip()]
            if len(source_candidates) >= 2:
                left_source = _resolve_source_name(source_candidates[0], rows[0] if rows else {})
                right_source = _resolve_source_name(source_candidates[1], rows[0] if rows else {})
        if rows and (left_source not in rows[0] or right_source not in rows[0]):
            continue
        for row in transformed_rows:
            left_value = _coerce_numeric(row.get(left_source, ''))
            right_value = _coerce_numeric(row.get(right_source, ''))
            if left_value is None or right_value is None:
                continue
            if '%' in str(row.get(right_source, '')):
                right_value = right_value / 100.0
            if operation == 'divide':
                if right_value == 0:
                    continue
                result = left_value / right_value
            elif operation == 'add':
                result = left_value + right_value
            elif operation == 'subtract':
                result = left_value - right_value
            else:
                result = left_value * right_value
            row[target] = _format_numeric_like_source(result, row.get(target, row.get(left_source, '')))
    return transformed_rows


def _generate_derived_value(column_name: str, instruction: str, row_index: int) -> str:
    fake = Faker()
    fake.seed_instance(1000 + row_index)
    hint = f"{column_name} {instruction}".lower()
    if 'social security' in hint or 'ssn' in hint:
        return fake.ssn()
    if 'email' in hint:
        return fake.email()
    if 'phone' in hint:
        return fake.phone_number()
    if 'address' in hint:
        return fake.street_address()
    if 'city' in hint:
        return fake.city()
    if 'state' in hint:
        return fake.state_abbr()
    if 'zip' in hint or 'postal' in hint:
        return fake.postcode()
    if 'first name' in hint:
        return fake.first_name()
    if 'last name' in hint:
        return fake.last_name()
    if 'full name' in hint or hint.strip().endswith('name'):
        return fake.name()
    if 'routing' in hint:
        return ''.join(str(fake.random_digit()) for _ in range(9))
    if 'account' in hint or 'member id' in hint or 'employee id' in hint or 'subscriber id' in hint:
        return ''.join(str(fake.random_digit()) for _ in range(8))
    if 'dob' in hint or 'date of birth' in hint or 'birth' in hint:
        return fake.date_of_birth(minimum_age=21, maximum_age=70).strftime('%m/%d/%Y')
    return instruction


def _apply_structured_output_plan(rows: list[dict[str, Any]], output_plan: list[dict[str, Any]] | None, lookup_rows: list[dict[str, Any]] | None = None) -> tuple[list[str], list[dict[str, Any]]] | None:
    if not rows or not output_plan:
        return None
    output_headers = [item.get('name') or '' for item in output_plan if (item.get('name') or '').strip()]
    if not output_headers:
        return None
    transformed_rows = _filter_visible_rows(
        _apply_output_plan_lookup(_apply_output_plan_arithmetic(rows, output_plan), output_plan, lookup_rows),
        output_plan,
    )
    built_rows: list[dict[str, Any]] = []
    for row_index, row in enumerate(transformed_rows):
        out: dict[str, Any] = {}
        for item in output_plan:
            name = (item.get('name') or '').strip()
            if not name:
                continue
            operation = (item.get('operation') or 'keep').strip().lower()
            source_a = _resolve_source_name(item.get('source_a') or '', row)
            notes = (item.get('instructions') or '').strip()
            if operation in {'keep', 'rename'} and source_a and source_a in row:
                out[name] = row.get(source_a, '')
            elif operation in {'multiply', 'divide', 'add', 'subtract'} and name in row:
                out[name] = row.get(name, '')
            elif source_a and source_a in row:
                out[name] = row.get(source_a, '')
            else:
                out[name] = _generate_derived_value(name, notes, row_index)
        built_rows.append(out)
    return output_headers, built_rows


def _apply_hidden_options(rows: list[dict[str, Any]], source_headers: list[str] | None = None, hidden_row_indexes: list[int] | None = None, hidden_columns: list[str] | None = None, ignore_hidden_rows: bool = False, ignore_hidden_columns: bool = False) -> tuple[list[str] | None, list[dict[str, Any]]]:
    filtered_rows = list(rows)
    filtered_headers = list(source_headers) if source_headers is not None else None
    if ignore_hidden_rows and hidden_row_indexes:
        hidden_set = set(hidden_row_indexes)
        filtered_rows = [row for idx, row in enumerate(filtered_rows) if idx not in hidden_set]
    if ignore_hidden_columns and hidden_columns:
        hidden_set = set(hidden_columns)
        if filtered_headers is not None:
            filtered_headers = [header for header in filtered_headers if header not in hidden_set]
        filtered_rows = [{k: v for k, v in row.items() if k not in hidden_set} for row in filtered_rows]
    return filtered_headers, filtered_rows


def apply_transform_plan(*, rows: list[dict[str, Any]], plan: dict[str, Any], output_plan: list[dict[str, Any]] | None = None, source_headers: list[str] | None = None, hidden_row_indexes: list[int] | None = None, hidden_columns: list[str] | None = None, ignore_hidden_rows: bool = False, ignore_hidden_columns: bool = False, lookup_rows: list[dict[str, Any]] | None = None) -> tuple[list[str], list[dict[str, Any]]]:
    _, rows = _apply_hidden_options(rows, source_headers=source_headers, hidden_row_indexes=hidden_row_indexes, hidden_columns=hidden_columns, ignore_hidden_rows=ignore_hidden_rows, ignore_hidden_columns=ignore_hidden_columns)
    structured_result = _apply_structured_output_plan(rows, output_plan, lookup_rows=lookup_rows)
    if structured_result is not None:
        return structured_result

    output_columns = plan.get('output_columns') or []
    filters = plan.get('filters') or []
    if not output_columns:
        raise SpreadsheetTransformError('Transform plan did not include output columns.')

    rows = _apply_output_plan_arithmetic(rows, output_plan)
    filtered_rows = []
    for row in rows:
        keep = True
        for rule in filters:
            column = (rule.get('column') or '').strip()
            operator = (rule.get('operator') or '').strip()
            value = str(rule.get('value') or '')
            cell = str(row.get(column, '') or '')
            if operator == 'equals' and cell != value:
                keep = False
            elif operator == 'not_equals' and cell == value:
                keep = False
            elif operator == 'contains' and value.lower() not in cell.lower():
                keep = False
        if keep:
            filtered_rows.append(row)

    headers = [col.get('name') or 'Column' for col in output_columns]
    transformed = []
    for row_index, row in enumerate(filtered_rows):
        out = {}
        for col in output_columns:
            name = col.get('name') or 'Column'
            source = col.get('source')
            instruction = (col.get('instruction') or '').strip()
            if source:
                out[name] = row.get(source, '')
            else:
                out[name] = _generate_derived_value(name, instruction, row_index)
        transformed.append(out)
    return headers, transformed


def export_transform_csv(headers: list[str], rows: list[dict[str, Any]]) -> bytes:
    buffer = StringIO()
    writer = csv.DictWriter(buffer, fieldnames=headers)
    writer.writeheader()
    for row in rows:
        writer.writerow({header: row.get(header, '') for header in headers})
    return buffer.getvalue().encode('utf-8')


def _xlsx_number_format_for_column(column_name: str, output_plan: list[dict[str, Any]] | None) -> str | None:
    if not output_plan:
        return None
    for item in output_plan:
        if (item.get('name') or '').strip() != column_name:
            continue
        fmt = (item.get('format') or 'keep_source').strip().lower()
        if fmt == 'number' or fmt == 'decimal_hours':
            return '0.00'
        if fmt == 'currency':
            return '$#,##0.00'
        if fmt == 'percent':
            return '0.00%'
        if fmt == 'text':
            return '@'
        return None
    return None


def _xlsx_typed_value(value: Any, column_name: str, output_plan: list[dict[str, Any]] | None):
    text = value if isinstance(value, str) else str(value if value is not None else '')
    fmt = _xlsx_number_format_for_column(column_name, output_plan)
    if fmt == '@':
        return text
    numeric = _coerce_numeric(text)
    if numeric is None:
        return text
    if fmt == '0.00%':
        if '%' in text:
            return numeric / 100.0
        return numeric
    return numeric


def export_transform_xlsx(headers: list[str], rows: list[dict[str, Any]], output_plan: list[dict[str, Any]] | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Transformed'
    ws.append(headers)
    for row in rows:
        ws.append([_xlsx_typed_value(row.get(header, ''), header, output_plan) for header in headers])
    for col_idx, header in enumerate(headers, start=1):
        number_format = _xlsx_number_format_for_column(header, output_plan)
        if not number_format:
            continue
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = number_format
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
